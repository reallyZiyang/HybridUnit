#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Utility for editing Luban Excel config workbooks in this project.

Run this script as a UTF-8 file instead of piping Python through PowerShell.
That avoids corrupting Chinese workbook names such as B-Battle_战斗表.xlsx.

Examples:
  python Excels/Tools/ConfigWorkbookTool.py inspect
  python Excels/Tools/ConfigWorkbookTool.py apply-battle
  python Excels/Tools/ConfigWorkbookTool.py generate
"""

from __future__ import annotations

import argparse
import json
import subprocess
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[2]
EXCEL_ROOT = PROJECT_ROOT / "Excels"
DATA_ROOT = EXCEL_ROOT / "Excels"

TABLES_PATH = DATA_ROOT / "__tables__.xlsx"
BEANS_PATH = DATA_ROOT / "__beans__.xlsx"
ENUMS_PATH = DATA_ROOT / "__enums__.xlsx"
BATTLE_WORKBOOK_PATH = DATA_ROOT / "B-Battle_战斗表.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TYPE_FILL = PatternFill("solid", fgColor="E2F0D9")
GROUP_FILL = PatternFill("solid", fgColor="FFF2CC")
COMMENT_FILL = PatternFill("solid", fgColor="FCE4D6")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


def last_nonempty_row(ws) -> int:
    last = 0
    for row in range(1, ws.max_row + 1):
        if any(ws.cell(row, col).value is not None for col in range(1, ws.max_column + 1)):
            last = row
    return last


def delete_rows(ws, rows: Iterable[int]) -> None:
    for row in sorted(set(rows), reverse=True):
        ws.delete_rows(row, 1)


def normalize_boolean_formula_cells(ws, columns: Iterable[int], start_row: int = 4) -> None:
    for row in range(start_row, ws.max_row + 1):
        for column in columns:
            value = ws.cell(row, column).value
            if isinstance(value, str):
                normalized = value.strip().upper()
                if normalized in {"=TRUE()", "TRUE()"}:
                    ws.cell(row, column, True)
                elif normalized in {"=FALSE()", "FALSE()"}:
                    ws.cell(row, column, False)


def style_data_sheet(ws) -> None:
    for row_idx, fill in [(1, HEADER_FILL), (2, TYPE_FILL), (3, GROUP_FILL), (4, COMMENT_FILL)]:
        if row_idx > ws.max_row:
            continue

        for cell in ws[row_idx]:
            cell.fill = fill
            cell.alignment = CENTER
            if row_idx == 1:
                cell.font = BOLD

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws, max_width=64)


def ensure_damage_effect_hit_reaction() -> None:
    wb = openpyxl.load_workbook(BATTLE_WORKBOOK_PATH)
    ws = wb["DamageEffect"]

    header_by_name = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    can_crit_col = header_by_name.get("canCrit")
    hit_count_col = header_by_name.get("hitCount")
    reaction_col = header_by_name.get("playHitReaction")

    if reaction_col is None:
        insert_at = (hit_count_col or ws.max_column) + 1
        ws.insert_cols(insert_at)
        reaction_col = insert_at
        ws.cell(1, reaction_col, "playHitReaction")
        ws.cell(2, reaction_col, "bool")
        ws.cell(3, reaction_col, "c")
        ws.cell(4, reaction_col, "是否播放受击")

    id_col = header_by_name.get("id", 2)
    for row in range(5, ws.max_row + 1):
        effect_id = ws.cell(row, id_col).value
        if effect_id is None:
            continue

        if can_crit_col is not None:
            ws.cell(row, can_crit_col, effect_id in {3001, 3002})
        ws.cell(row, reaction_col, effect_id != 3003)

    style_data_sheet(ws)
    wb.save(BATTLE_WORKBOOK_PATH)
    print(f"Updated DamageEffect.playHitReaction: {BATTLE_WORKBOOK_PATH}")


def style_enum_sheet(ws) -> None:
    if "H1:L1" not in [str(cell_range) for cell_range in ws.merged_cells.ranges]:
        ws.merge_cells("H1:L1")

    for row_idx, fill in [(1, HEADER_FILL), (2, TYPE_FILL), (3, COMMENT_FILL)]:
        for cell in ws[row_idx]:
            cell.fill = fill
            cell.alignment = CENTER
            if row_idx == 1:
                cell.font = BOLD

    ws.freeze_panes = "A4"
    autofit_columns(ws, max_width=44)


def autofit_columns(ws, max_width: int) -> None:
    for col in range(1, ws.max_column + 1):
        width = 10
        for cell in ws[get_column_letter(col)]:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width


def replace_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        header_width = len(rows[0]) if rows else 0
        for row_index, row in enumerate(rows, start=1):
            if row_index >= 5 and header_width > 0 and len(row) == header_width - 1 and row[0] is not None:
                ws.append([None, *row])
            else:
                ws.append(row)
        style_data_sheet(ws)

    wb.save(path)


def resolve_project_path(value: str) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return PROJECT_ROOT / path


def upsert_table_rows(rows: list[list[object]]) -> None:
    wb = openpyxl.load_workbook(TABLES_PATH)
    ws = wb.active
    normalize_boolean_formula_cells(ws, columns=[4])
    full_names = {row[1] for row in rows}
    delete_rows(ws, (row for row in range(4, ws.max_row + 1) if ws.cell(row, 2).value in full_names))

    start = last_nonempty_row(ws) + 2
    for row_index, row_values in enumerate(rows, start=start):
        for col_index, value in enumerate(row_values, start=1):
            ws.cell(row_index, col_index, value)

    wb.save(TABLES_PATH)


def remove_bean_blocks(ws, full_names: set[str]) -> None:
    rows_to_delete: list[int] = []
    row = 4
    while row <= ws.max_row:
        full_name = ws.cell(row, 2).value
        if full_name in full_names:
            rows_to_delete.append(row)
            next_row = row + 1
            while next_row <= ws.max_row and ws.cell(next_row, 2).value is None:
                has_value = any(ws.cell(next_row, col).value is not None for col in range(1, ws.max_column + 1))
                if not has_value:
                    break
                rows_to_delete.append(next_row)
                next_row += 1
            row = next_row
        else:
            row += 1

    delete_rows(ws, rows_to_delete)


def upsert_bean_blocks(blocks: dict[str, list[tuple[str, str]]]) -> None:
    wb = openpyxl.load_workbook(BEANS_PATH)
    ws = wb.active
    remove_bean_blocks(ws, set(blocks))

    row_index = last_nonempty_row(ws) + 2
    for full_name, fields in blocks.items():
        for field_index, (field_name, field_type) in enumerate(fields):
            if field_index == 0:
                row_values = [None, full_name, None, None, ",", None, bean_comment(full_name), None, None, field_name, None, field_type]
            else:
                row_values = [None, None, None, None, None, None, None, None, None, field_name, None, field_type]

            for col_index, value in enumerate(row_values, start=1):
                ws.cell(row_index, col_index, value)
            row_index += 1

        row_index += 1

    wb.save(BEANS_PATH)


def bean_comment(full_name: str) -> str:
    comments = {
        "Battle.EffectRef": "效果引用",
        "Battle.BattleShapeDesc": "战斗范围形状",
    }
    return comments.get(full_name, full_name)


def replace_enum_sheet(sheet_name: str, enum_defs: list[tuple[str, str, list[tuple[str, str, int, str]]]]) -> None:
    wb = openpyxl.load_workbook(ENUMS_PATH)
    for existing_ws in wb.worksheets:
        normalize_boolean_formula_cells(existing_ws, columns=[3, 4])

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.append(["##var", "full_name", "flags", "unique", "group", "comment", "tags", "*items", None, None, None, None])
    ws.append(["##var", None, None, None, None, None, None, "name", "alias", "value", "comment", "tags"])
    ws.append(["##", "全名(包含模块和名字)", "是否为位标记枚举", "枚举项是否唯一", None, None, None, "枚举名", "别名", "值", "注释", None])

    for full_name, comment, items in enum_defs:
        first = True
        for name, alias, value, item_comment in items:
            if first:
                ws.append([None, full_name, False, True, None, comment, None, name, alias, value, item_comment, None])
                first = False
            else:
                ws.append([None, None, None, None, None, None, None, name, alias, value, item_comment, None])
        ws.append([None] * 12)

    style_enum_sheet(ws)
    wb.save(ENUMS_PATH)


def battle_workbook_sheets() -> dict[str, list[list[object]]]:
    return {
        "Unit": [
            ["##var", "id", "name", "radius", "camp", "layer", "attrs", "defaultSkills", "renderKey"],
            ["##type", "int", "string", "float", "int", "int", "(array#sep=|),(Attr.AttributePair#sep=,)", "(array#sep=|),int", "string"],
            ["##group", "c", "c", "c", "c", "c", "c", "c", "c"],
            ["##", "单位id", "单位名", "碰撞半径", "默认阵营", "层级", "初始属性", "默认技能", "表现资源key"],
            [1001, "玩家", 0.35, 1, 0, "102,1000|105,1000|101,50|103,10|104,300", "2001|2002", "player"],
            [1101, "小怪", 0.35, 2, 0, "102,300|105,300|101,20|103,3|104,220", "2101", "monster_melee"],
        ],
        "Skill": [
            ["##var", "id", "name", "actionName", "castPreMs", "castBackMs", "cooldownMs", "targetType", "selectType", "shape", "effects"],
            ["##type", "int", "string", "string", "int", "int", "int", "Battle.SkillTargetType", "Battle.TargetSelectType", "Battle.BattleShapeDesc", "(array#sep=|),Battle.EffectRef"],
            ["##group", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c"],
            ["##", "技能id", "技能名", "动作名", "前摇毫秒", "后摇毫秒", "冷却毫秒", "目标类型", "选敌方式", "范围形状", "效果列表"],
            [2001, "普攻", "attack", 300, 200, 1000, "Enemy", "Nearest", "Circle,1.2,0,0,0,0,0", "Damage,3001,10000"],
            [2002, "火球术", "skill_1", 400, 300, 2000, "Enemy", "Nearest", "Null,0,0,0,0,0,0", "Projectile,6001,0"],
            [2101, "怪物撕咬", "attack", 500, 200, 1500, "Enemy", "Nearest", "Circle,1,0,0,0,0,0", "Damage,3001,8000"],
        ],
        "Buff": [
            ["##var", "id", "name", "durationMs", "maxStack", "stackMode", "tickMs", "attrs", "tickEffects", "beginEffects", "endEffects"],
            ["##type", "int", "string", "int", "int", "Battle.BuffStackMode", "int", "(array#sep=|),(Attr.AttributePair#sep=,)", "(array#sep=|),Battle.EffectRef", "(array#sep=|),Battle.EffectRef", "(array#sep=|),Battle.EffectRef"],
            ["##group", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c"],
            ["##", "BuffId", "Buff名", "持续毫秒", "最大层数", "叠加规则", "周期毫秒", "属性修改", "周期效果", "开始效果", "结束效果"],
            [7001, "燃烧", 5000, 3, "RefreshAndStack", 1000, "", "Damage,3003,10000", "", ""],
            [7002, "加速", 3000, 1, "Refresh", 0, "104,500", "", "", ""],
        ],
        "DamageEffect": [
            ["##var", "id", "name", "attr", "ratio", "fixedValue", "damageElement", "canCrit", "hitCount", "playHitReaction"],
            ["##type", "int", "string", "Attr.AttributeType", "long", "long", "int", "bool", "int", "bool"],
            ["##group", "c", "c", "c", "c", "c", "c", "c", "c", "c"],
            ["##", "效果id", "效果名", "取值属性", "属性倍率", "固定伤害", "伤害元素", "是否暴击", "最大命中数", "是否播放受击"],
            [3001, "普攻伤害", "Atk", 10000, 0, 0, True, 1, True],
            [3002, "火球命中伤害", "Atk", 12000, 0, 0, True, 1, True],
            [3003, "燃烧周期伤害", "Atk", 3000, 0, 1, False, 0, False],
        ],
        "HealEffect": [
            ["##var", "id", "name", "attr", "ratio", "fixedValue", "canCrit"],
            ["##type", "int", "string", "Attr.AttributeType", "long", "long", "bool"],
            ["##group", "c", "c", "c", "c", "c", "c"],
            ["##", "效果id", "效果名", "取值属性", "属性倍率", "固定回血", "是否暴击"],
            [4001, "小治疗", "HpMax", 1000, 0, False],
            [4002, "固定回血", "Null", 0, 50, False],
        ],
        "AddBuffEffect": [
            ["##var", "id", "name", "buffId", "durationOverrideMs", "stack"],
            ["##type", "int", "string", "int", "int", "int"],
            ["##group", "c", "c", "c", "c", "c"],
            ["##", "效果id", "效果名", "BuffId", "覆盖持续时间毫秒", "增加层数"],
            [5001, "添加燃烧", 7001, 0, 1],
            [5002, "添加加速", 7002, 3000, 1],
        ],
        "ProjectileEffect": [
            ["##var", "id", "name", "projectileKey", "speed", "radius", "lifetimeMs", "pierceCount", "hitIntervalMs", "hitEffects"],
            ["##type", "int", "string", "string", "float", "float", "int", "int", "int", "(array#sep=|),Battle.EffectRef"],
            ["##group", "c", "c", "c", "c", "c", "c", "c", "c", "c"],
            ["##", "效果id", "效果名", "抛射物表现key", "速度", "碰撞半径", "存活毫秒", "穿透次数", "同目标命中间隔", "命中效果"],
            [6001, "火球", "projectile_fireball", 8.0, 0.25, 3000, 1, 0, "Damage,3002,10000|AddBuff,5001,0"],
        ],
    }


def apply_battle() -> None:
    replace_workbook(BATTLE_WORKBOOK_PATH, battle_workbook_sheets())
    upsert_table_rows(
        [
            ["战斗单位表", "Battle.TbUnit", "Battle.UnitCfg", True, "Unit@B-Battle_战斗表.xlsx", None, None, "c", None, None, None],
            ["技能表", "Battle.TbSkill", "Battle.SkillCfg", True, "Skill@B-Battle_战斗表.xlsx", None, None, "c", None, None, None],
            ["Buff表", "Battle.TbBuff", "Battle.BuffCfg", True, "Buff@B-Battle_战斗表.xlsx", None, None, "c", None, None, None],
            ["伤害效果表", "Battle.TbDamageEffect", "Battle.DamageEffectCfg", True, "DamageEffect@B-Battle_战斗表.xlsx", None, None, "c", None, None, None],
            ["回血效果表", "Battle.TbHealEffect", "Battle.HealEffectCfg", True, "HealEffect@B-Battle_战斗表.xlsx", None, None, "c", None, None, None],
            ["添加Buff效果表", "Battle.TbAddBuffEffect", "Battle.AddBuffEffectCfg", True, "AddBuffEffect@B-Battle_战斗表.xlsx", None, None, "c", None, None, None],
            ["抛射物效果表", "Battle.TbProjectileEffect", "Battle.ProjectileEffectCfg", True, "ProjectileEffect@B-Battle_战斗表.xlsx", None, None, "c", None, None, None],
        ]
    )
    upsert_bean_blocks(
        {
            "Battle.EffectRef": [
                ("type", "Battle.EffectType"),
                ("id", "int"),
                ("value", "long"),
            ],
            "Battle.BattleShapeDesc": [
                ("shapeType", "Battle.SkillShapeType"),
                ("radius", "float"),
                ("width", "float"),
                ("length", "float"),
                ("angleDeg", "float"),
                ("offsetX", "float"),
                ("offsetY", "float"),
            ],
        }
    )
    replace_enum_sheet(
        "战斗",
        [
            (
                "Battle.EffectType",
                "效果类型",
                [
                    ("Null", "无", 0, "无"),
                    ("Damage", "伤害", 1, "造成伤害"),
                    ("Heal", "回血", 2, "回复生命"),
                    ("AddBuff", "添加Buff", 3, "添加Buff"),
                    ("Projectile", "抛射物", 4, "发射抛射物"),
                ],
            ),
            (
                "Battle.SkillTargetType",
                "技能目标类型",
                [
                    ("Null", "无", 0, "无"),
                    ("Self", "自身", 1, "自身"),
                    ("Enemy", "敌方", 2, "敌方"),
                    ("Ally", "友方", 3, "友方"),
                    ("Position", "位置", 4, "指定位置"),
                    ("Direction", "方向", 5, "指定方向"),
                ],
            ),
            (
                "Battle.TargetSelectType",
                "目标选择类型",
                [
                    ("Null", "无", 0, "无"),
                    ("Nearest", "最近", 1, "最近目标"),
                    ("Random", "随机", 2, "随机目标"),
                    ("LowestHp", "最低血量", 3, "最低血量目标"),
                    ("SelfAround", "自身周围", 4, "以自身为中心"),
                    ("Forward", "前方", 5, "朝向前方"),
                ],
            ),
            (
                "Battle.SkillShapeType",
                "技能形状类型",
                [
                    ("Null", "无", 0, "无"),
                    ("Circle", "圆形", 1, "圆形"),
                    ("Rect", "矩形", 2, "矩形"),
                    ("Sector", "扇形", 3, "扇形"),
                    ("Capsule", "胶囊", 4, "胶囊线段"),
                ],
            ),
            (
                "Battle.BuffStackMode",
                "Buff叠加规则",
                [
                    ("Refresh", "刷新", 1, "刷新持续时间"),
                    ("Stack", "叠层", 2, "只增加层数"),
                    ("RefreshAndStack", "刷新并叠层", 3, "刷新时间并增加层数"),
                    ("Replace", "替换", 4, "替换旧Buff"),
                ],
            ),
        ],
    )
    print(f"Applied Battle workbook schema: {BATTLE_WORKBOOK_PATH}")


def apply_spec(spec_path: Path) -> None:
    spec = json.loads(spec_path.read_text(encoding="utf-8"))

    for workbook in spec.get("workbooks", []):
        replace_workbook(resolve_project_path(workbook["path"]), workbook["sheets"])

    if "tables" in spec:
        upsert_table_rows(spec["tables"])

    if "beans" in spec:
        upsert_bean_blocks(spec["beans"])

    for sheet_name, enum_defs in spec.get("enumSheets", {}).items():
        replace_enum_sheet(sheet_name, enum_defs)

    print(f"Applied config workbook spec: {spec_path}")


def inspect() -> None:
    for path in [TABLES_PATH, BEANS_PATH, ENUMS_PATH, BATTLE_WORKBOOK_PATH]:
        print(f"\nFILE {path.relative_to(PROJECT_ROOT)}")
        if not path.exists():
            print("  missing")
            continue

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            print(f"  SHEET {ws.title} rows={ws.max_row} cols={ws.max_column}")
            for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_index > 8:
                    break
                print("   ", row[: min(12, len(row))])


def generate() -> None:
    script = EXCEL_ROOT / "Gen_Codex.bat"
    completed = subprocess.run(["cmd", "/c", str(script)], cwd=PROJECT_ROOT, check=False)
    raise SystemExit(completed.returncode)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Edit and generate Luban config workbooks.")
    parser.add_argument("command", choices=["inspect", "apply-battle", "apply-battle-hit-reaction", "apply-spec", "generate"])
    parser.add_argument("spec", nargs="?", help="UTF-8 JSON spec path for apply-spec.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.command == "inspect":
        inspect()
    elif args.command == "apply-battle":
        apply_battle()
    elif args.command == "apply-battle-hit-reaction":
        ensure_damage_effect_hit_reaction()
    elif args.command == "apply-spec":
        if not args.spec:
            raise SystemExit("apply-spec requires a spec path")
        apply_spec(resolve_project_path(args.spec))
    elif args.command == "generate":
        generate()


if __name__ == "__main__":
    main()
